import serial
import serial.tools.list_ports
import re
import openpyxl as pyxl
import datetime
import time
import threading


class scaleRead:
    _DEBUG = False
    running = False
    weight = -1

    read_t = None

    def __init__(self, serial_port, debug=False) -> None:
        self._DEBUG = debug
        if not self._DEBUG:
            self.ser = serial.Serial(serial_port, 9600, timeout=1)

    @staticmethod
    def detect_ports():
        return [x.device for x in serial.tools.list_ports.comports()]

    def start(self):
        if not self._DEBUG:
            self.read_t = threading.Thread(target=self._reading_thread)
            self.running = True
            self.read_t.start()

    def _reading_thread(self):
        while self.running:
            # print("rt", self.running)
            response = str(self.ser.read_until(b"\n\r\n\n\n"))
            self.weight = float(re.findall(r'\d+\.\d+', response)[0])

    def read(self) -> float:
        if self._DEBUG:
            return round(time.time(), 2)
        else:
            return self.weight

    @staticmethod
    def write_to_file(self, weight: float, fp: str) -> None:
        wb = pyxl.Workbook()
        ws = wb.active

        ws['A1'] = "Index"
        ws['A2'] = 2
        ws['B1'] = "Time"
        ws['B2'] = datetime.datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")
        ws['C1'] = "Weight"
        ws['C2'] = round(weight, 2)
        ws['D1'] = "Unit"
        ws['D2'] = "g"
        wb.save(fp)

    def stop(self):
        if self._DEBUG:
            return
        self.running = False
        self.read_t.join()

    def __del__(self):
        print("DEL")
        if not self._DEBUG:
            self.ser.close()


if __name__ == "__main__":
    sr = scaleRead("COM11")
    sr.start()

    for i in range(10):
        weight = sr.read()
        print(weight)
        time.sleep(0.514)

    sr.stop()
    print("_DEL")
    del sr
